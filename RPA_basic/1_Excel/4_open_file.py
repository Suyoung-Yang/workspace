from openpyxl import load_workbook # 파일 불러오기 역할
wb = load_workbook("sample.xlsx") # 파일에서 워크북 불러오기
ws = wb.active # 활성화된 시트

# cell 데이터 불러오기
for x in range(1,11):
    for y in range(1,11):
        print(ws.cell(row=x,column = y).value,end=' ')
    print()

print("="*30)

# cell 개수를 모를 때
for x in range(1, ws.max_row):
    for y in range(1,ws.max_column):
        print(ws.cell(row=x,column= y).value,end=' ')
    print()
