import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.detach(), encoding = 'utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.detach(), encoding = 'utf-8')


from random import *
from openpyxl import Workbook
wb = Workbook()
ws = wb.active

# 1줄씩 데이터 넣기
ws.append(["번호","영어","수학"])
for i in range(1,11):
    ws.append([i,randint(0,100),randint(0,100)])

# 영어 점수 가져오기
col_B = ws["B"] # 컬럼은 A,B,C니까, 영어 컬럼만 가져오기
print(col_B) # (<Cell 'Sheet'.B1>, <Cell 'Sheet'.B2>, ... ,
for cell in col_B:
    print(cell.value)

col_range = ws["B:C"] # 슬라이싱, 영어 수학 점수 함께 가져오기

for col in col_range:
    for cell in col:
        print(cell.value)

# row도 가능함.

row_title = ws[1] # 1번째 row만 가져오기
for cell in row_title:
    print(cell.value)

print("-"*100)
row_range = ws[2:6] # 여기서 슬라이싱은 포함됨을 기억!!!!!!!!!!!
for rows in row_range:
    for cell in rows:
        print(cell.value,end=' ')
    print()

print("-"*100)

from openpyxl.utils.cell import coordinate_from_string



row_range = ws[2:ws.max_row] # 2번째부터 마지막줄까지
for rows in row_range:
    for cell in rows:
        # print(cell.value,end=' ') # 셀 값 가져오기
        # print(cell.coordinate,end=' ') # 셀에 대한 정보 가져오기
        xy = coordinate_from_string(cell.coordinate) # tuple 형태로 가져오기
        print(xy,end = ' ')
        print(xy[0],end=' ')
    print()




wb.save("sample.xlsx")
# https://www.youtube.com/watch?v=exgO1LFl9x8
# 52:02 까지 완료.
