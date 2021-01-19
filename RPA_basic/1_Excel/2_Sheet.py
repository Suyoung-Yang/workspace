from openpyxl import Workbook
wb = Workbook()
ws = wb.create_sheet() # 새로운 시트를 기본 이름으로 생성
ws.title = "MySheet"
ws.sheet_properties.tabColor="ff0000" # # 없이 글자만 입력 rgb 형태로

ws1 = wb.create_sheet("YourSheet") # 주어진 이름으로 sheet 생성
ws2 = wb.create_sheet("NewSheet",2) # 2번 인덱스에 생성.

new_ws = wb["NewSheet"] # dict 형태로 시트에 접근이 가능하다.

print(wb.sheetnames) # 모든 시트 이름 확인 가능함.

# sheet 복사
new_ws["A1"] = "Test" # A1 셀에 입력.
target = wb.copy_worksheet(new_ws) # wb란 워크북(파일)에서 new_ws 탭을 복사해서 target이란 탭을 만든다.
target.title = "Copied sheet" # target의 탭의 title은 다음과 같다.






wb.save("sample.xlsx")
